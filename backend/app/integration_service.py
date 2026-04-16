"""
Integration service — Template-driven Excel integration with approval workflow.
"""
from io import BytesIO
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from app.models import IntegrationUpload, IntegrationItem, IntegrationApproval

# Template column mapping (1-indexed → field)
TEMPLATE_MAP = {
    1: "level", 2: "title", 3: "revision", 4: "montaj_no", 5: "quantity",
    6: "description", 7: "maturity_state", 8: "owner", 9: "dokuman_malzeme_turu",
    10: "siparis_raw", 11: "birim", 13: "miktar", 14: "df_tr",
    15: "kalem_tipi", 16: "m_turu", 18: "malzeme_no", 19: "sap_usage",
    20: "kullanim_miktari", 21: "ana_malzeme",
}


def _apply_siparis_durumu(rows: list[dict]):
    l1_f_active = False
    l2_f_active = False
    l3_trigger = ""

    for r in rows:
        level = r.get("level", 0)
        kt = (r.get("kalem_tipi") or "").strip().upper()

        if level == 1:
            l1_f_active = kt == "F"
            r["siparis_durumu"] = "EVET" if l1_f_active else "HAYIR"
        elif level == 2:
            l3_trigger = ""
            l2_f_active = kt == "F"
            r["siparis_durumu"] = "EVET" if l2_f_active else ("EVET" if l1_f_active else "HAYIR")
        elif level == 3:
            if kt == "F":
                l3_trigger = "F"
                r["siparis_durumu"] = "EVET"
            elif kt == "Y":
                l3_trigger = "Y"
                r["siparis_durumu"] = "EVET" if l2_f_active else "HAYIR"
            elif kt in ("H", "C"):
                l3_trigger = ""
                r["siparis_durumu"] = "HAYIR"
            elif kt == "E":
                l3_trigger = ""
                r["siparis_durumu"] = "EVET"
            else:
                l3_trigger = ""
                r["siparis_durumu"] = "EVET" if l2_f_active else "HAYIR"
        else:  # level 4+
            if l3_trigger == "F":
                r["siparis_durumu"] = "EVET"
            elif l3_trigger == "Y":
                r["siparis_durumu"] = "HAYIR"
            else:
                r["siparis_durumu"] = "EVET" if l2_f_active else "HAYIR"


def _apply_montaj_flag(rows: list[dict]):
    for r in rows:
        kt = (r.get("kalem_tipi") or "").strip().upper()
        r["montaj_mi"] = "EVET" if kt == "F" else "HAYIR"


def _apply_quantity_calculation(rows: list[dict]):
    parent_stack = []
    for r in rows:
        level = r.get("level", 0)
        qty = float(r.get("quantity") or 1.0)

        while parent_stack and parent_stack[-1]["level"] >= level:
            parent_stack.pop()

        kull_raw = r.get("kullanim_miktari")
        kull = None
        if kull_raw:
            try:
                kull = float(str(kull_raw).replace(",", ".").strip())
            except (ValueError, TypeError):
                kull = None

        if kull and kull > 0:
            montaj_qty = 1.0
            for p in parent_stack:
                montaj_qty *= p["quantity"]
            r["hesaplanan_miktar"] = kull * qty * montaj_qty
        else:
            r["hesaplanan_miktar"] = None

        parent_stack.append({"level": level, "quantity": qty})


async def parse_integration_file(
    filepath: str, upload: IntegrationUpload, db: AsyncSession, calculate_quantity: bool = False
):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        r = {}
        for col_idx, field in TEMPLATE_MAP.items():
            idx = col_idx - 1
            r[field] = vals[idx] if idx < len(vals) else None
        r["level"] = int(r.get("level") or 0)
        r["quantity"] = float(r.get("quantity") or 1.0) if r.get("quantity") else 1.0
        rows.append(r)

    _apply_siparis_durumu(rows)
    _apply_montaj_flag(rows)
    if calculate_quantity:
        _apply_quantity_calculation(rows)

    items = []
    for idx, r in enumerate(rows, 1):
        items.append(IntegrationItem(
            upload_id=upload.id,
            row_number=idx,
            level=r["level"],
            title=str(r.get("title") or "").strip(),
            revision=str(r.get("revision") or "").strip(),
            montaj_no=str(r.get("montaj_no") or "").strip(),
            quantity=r["quantity"],
            description=str(r.get("description") or "").strip(),
            maturity_state=str(r.get("maturity_state") or "").strip(),
            owner=str(r.get("owner") or "").strip(),
            dokuman_malzeme_turu=str(r.get("dokuman_malzeme_turu") or "").strip(),
            birim=str(r.get("birim") or "").strip(),
            miktar=float(r.get("miktar") or 0) if r.get("miktar") else None,
            df_tr=str(r.get("df_tr") or "").strip(),
            kalem_tipi=str(r.get("kalem_tipi") or "").strip(),
            m_turu=str(r.get("m_turu") or "").strip(),
            malzeme_no=str(r.get("malzeme_no") or "").strip(),
            sap_usage=str(r.get("sap_usage") or "").strip(),
            kullanim_miktari=str(r.get("kullanim_miktari") or "").strip(),
            ana_malzeme=str(r.get("ana_malzeme") or "").strip(),
            siparis_durumu=r.get("siparis_durumu", ""),
            montaj_mi=r.get("montaj_mi", ""),
            hesaplanan_miktar=r.get("hesaplanan_miktar"),
        ))

    db.add_all(items)
    upload.total_rows = len(items)
    upload.status = "processed"
    await db.commit()
    wb.close()
    return {"total": len(items)}


async def export_filtered_excel(upload_id: int, db: AsyncSession, **filters) -> BytesIO:
    q = select(IntegrationItem).where(IntegrationItem.upload_id == upload_id)

    if filters.get("siparis_durumu"):
        q = q.where(IntegrationItem.siparis_durumu == filters["siparis_durumu"])
    if filters.get("montaj_mi"):
        q = q.where(IntegrationItem.montaj_mi == filters["montaj_mi"])
    if filters.get("uzmanlik"):
        q = q.where(IntegrationItem.uzmanlik == filters["uzmanlik"])
    if filters.get("kalem_tipi"):
        q = q.where(IntegrationItem.kalem_tipi == filters["kalem_tipi"])
    if filters.get("level") is not None and filters["level"] != "":
        q = q.where(IntegrationItem.level == int(filters["level"]))

    q = q.order_by(IntegrationItem.row_number)
    result = await db.execute(q)
    items = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Integration Export"

    headers = [
        "#", "Level", "Title", "Montaj No", "Qty", "Kalem Tipi",
        "Sipariş Durumu", "Montaj mı?", "Uzmanlık", "Birim",
        "MalzemeNo", "Hes.Miktar", "DF TR", "Onay",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    evet_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    hayir_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for ri, item in enumerate(items, 2):
        vals = [
            item.row_number, item.level, item.title, item.montaj_no, item.quantity,
            item.kalem_tipi, item.siparis_durumu, item.montaj_mi, item.uzmanlik,
            item.birim, item.malzeme_no, item.hesaplanan_miktar, item.df_tr,
            "✓" if item.approved else "",
        ]
        fill = evet_fill if item.siparis_durumu == "EVET" else (hayir_fill if item.siparis_durumu == "HAYIR" else None)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Log download
    db.add(IntegrationApproval(
        upload_id=upload_id,
        action="download",
        filter_criteria=str(filters),
        details=f"{len(items)} satır indirildi",
    ))
    await db.commit()

    return buf


async def compare_reupload(upload_id: int, filepath: str, db: AsyncSession) -> dict:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Load existing items
    result = await db.execute(
        select(IntegrationItem).where(IntegrationItem.upload_id == upload_id).order_by(IntegrationItem.row_number)
    )
    existing = {i.row_number: i for i in result.scalars().all()}

    diffs = []
    approved_count = 0
    row_idx = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_idx += 1
        vals = list(row)
        if row_idx not in existing:
            continue

        item = existing[row_idx]
        if item.locked:
            continue

        # Check approval column (last column with ✓)
        approval_val = str(vals[-1] or "").strip() if len(vals) > 0 else ""
        is_approved = approval_val in ("✓", "✔", "x", "X", "1", "evet", "EVET")

        # Check field changes
        new_title = str(vals[2] or "").strip() if len(vals) > 2 else ""
        changes = {}
        if new_title and new_title != (item.title or ""):
            changes["title"] = {"old": item.title, "new": new_title}
            item.title = new_title

        if is_approved:
            item.approved = True
            item.approved_at = datetime.now(timezone.utc)
            item.locked = True
            approved_count += 1

        if changes:
            diffs.append({"row": row_idx, "changes": changes, "approved": is_approved})

    db.add(IntegrationApproval(
        upload_id=upload_id,
        action="reupload",
        details=f"{len(diffs)} değişiklik, {approved_count} onay",
    ))
    await db.commit()
    wb.close()

    return {"total_compared": row_idx, "diffs": diffs, "approved_count": approved_count}
