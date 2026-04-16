"""
Excel service — BOM parse, Material Master import, reprocess, export.
"""
import os
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models import BomProject, BomItem, MaterialMaster
from app.rules_engine import apply_rules

PLM_COLS = [
    "Level", "Title", "Revision", "Quantity", "Description", "MaturityState", "Owner",
    "CatiaAciklama", "YedekParcaMi", "MalzemeStandartDokumani", "ToleransDokumani",
    "DokumanMalzemeTuru", "MalzemeNo", "IstatistikselProsesKontrol", "ParcaStandartDokumani",
    "SAP Usage", "YanmazlikParametresi", "Hacim", "BoyaKodu", "YuzeyAlani",
    "HomologasyonDokumani", "EmisyonFaktoru", "FinishingStandartDokumani", "ReferansResim",
    "Kutle", "Sertlik", "ProjeKodu", "KullanimMiktari", "AnaMalzemeGrubu",
    "IsilIslemDokumani", "AnaMalzeme",
]

COL_MAP = {
    0: "level", 1: "title", 2: "revision", 3: "quantity", 4: "description",
    5: "maturity_state", 6: "owner", 7: "catia_aciklama", 8: "yedek_parca_mi",
    9: "malzeme_standart_dokumani", 10: "tolerans_dokumani", 11: "dokuman_malzeme_turu",
    12: "malzeme_no", 13: "istatistiksel_proses_kontrol", 14: "parca_standart_dokumani",
    15: "sap_usage", 16: "yanmazlik_parametresi", 17: "hacim", 18: "boya_kodu",
    19: "yuzey_alani", 20: "homologasyon_dokumani", 21: "emisyon_faktoru",
    22: "finishing_standart_dokumani", 23: "referans_resim", 24: "kutle",
    25: "sertlik", 26: "proje_kodu", 27: "kullanim_miktari", 28: "ana_malzeme_grubu",
    29: "isil_islem_dokumani", 30: "ana_malzeme",
}

LEVEL_FILLS = {
    0: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    1: PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    2: PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
    3: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}


async def parse_bom_file(filepath: str, project: BomProject, db: AsyncSession):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Pass 1: collect material numbers
    material_nos = set()
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 31:
            vals.extend([None] * (31 - len(vals)))
        rows_data.append(vals)
        malz = str(vals[12] or "").strip()
        title = str(vals[1] or "").strip()
        ana = str(vals[30] or "").strip()
        if malz:
            material_nos.add(malz)
        if title:
            material_nos.add(title)
        if ana:
            material_nos.add(ana)

    # Batch lookup from MaterialMaster
    material_map = {}
    if material_nos:
        result = await db.execute(
            select(MaterialMaster).where(MaterialMaster.material_no.in_(material_nos))
        )
        for mat in result.scalars().all():
            material_map[mat.material_no] = mat

    # Pass 2: process rows
    level1_title = ""
    level2_title = ""
    parent_stack = []  # [(level, qty)]
    items = []
    row_num = 0

    for vals in rows_data:
        row_num += 1
        level = int(vals[0] or 0)
        title = str(vals[1] or "").strip()
        quantity = float(vals[3] or 1.0) if vals[3] else 1.0
        malzeme_no = str(vals[12] or "").strip()
        sap_usage = str(vals[15] or "").strip()
        kullanim_miktari = str(vals[27] or "").strip() if vals[27] else ""
        ana_malzeme = str(vals[30] or "").strip()

        # Update hierarchy
        if level == 1:
            level1_title = title
        if level == 2:
            level2_title = title

        # Parent stack for qty calculation
        while parent_stack and parent_stack[-1][0] >= level:
            parent_stack.pop()
        parent_qty_product = 1.0
        for _, pqty in parent_stack:
            parent_qty_product *= pqty

        # Material lookup: malzeme_no → title → ana_malzeme
        mat = material_map.get(malzeme_no) or material_map.get(title) or material_map.get(ana_malzeme)
        mat_kt = mat.kalem_tipi if mat else None
        mat_birim = mat.birim if mat else None

        # Apply rules
        derived = apply_rules(
            level=level, title=title, malzeme_no=malzeme_no, ana_malzeme=ana_malzeme,
            sap_usage=sap_usage, quantity=quantity, level1_title=level1_title,
            level2_title=level2_title, material_kalem_tipi=mat_kt,
            material_birim=mat_birim, parent_qty_product=parent_qty_product,
            parent_montaj_kt=None,
        )

        # Build item
        item_data = {"project_id": project.id, "row_number": row_num}
        for idx, field in COL_MAP.items():
            item_data[field] = str(vals[idx]).strip() if vals[idx] is not None else ""
        item_data["level"] = level
        item_data["quantity"] = quantity
        item_data["kullanim_miktari"] = kullanim_miktari

        # Derived
        item_data.update(derived)
        birlestirme = f"{derived['montaj']}|{title}|{derived['malzeme_no_sap']}"
        item_data["birlestirme"] = birlestirme
        item_data["ana_malzeme_derived"] = ana_malzeme

        items.append(BomItem(**item_data))
        parent_stack.append((level, quantity))

    # Bulk insert
    db.add_all(items)

    # Update project stats
    total = len(items)
    unresolved = sum(1 for i in items if i.needs_review)
    project.total_rows = total
    project.resolved_rows = total - unresolved
    project.unresolved_rows = unresolved
    project.status = "review" if unresolved > 0 else "completed"
    await db.commit()

    wb.close()
    return {"total": total, "unresolved": unresolved}


async def import_mm03_file(filepath: str, db: AsyncSession) -> dict:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Find header row
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            val = str(cell.value or "").lower().strip()
            if any(k in val for k in ("malzeme", "material", "numara", "no")):
                headers["material_no"] = cell.column - 1
            elif any(k in val for k in ("kalem", "item", "tipi", "category")):
                headers["kalem_tipi"] = cell.column - 1
            elif any(k in val for k in ("birim", "unit")):
                headers["birim"] = cell.column - 1
            elif any(k in val for k in ("aciklama", "desc", "açıklama")):
                headers["description"] = cell.column - 1
        if "material_no" in headers and "kalem_tipi" in headers:
            break

    if "material_no" not in headers:
        wb.close()
        return {"error": "MalzemeNo sütunu bulunamadı", "total": 0}

    imported = 0
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        mat_no = str(vals[headers["material_no"]] or "").strip()
        if not mat_no or mat_no.lower() in ("#n/a", "nan", "none"):
            continue
        kt = str(vals[headers.get("kalem_tipi", -1)] or "").strip() if "kalem_tipi" in headers else ""
        birim = str(vals[headers.get("birim", -1)] or "").strip() if "birim" in headers else ""
        desc = str(vals[headers.get("description", -1)] or "").strip() if "description" in headers else ""

        if not kt:
            continue

        existing = await db.execute(
            select(MaterialMaster).where(MaterialMaster.material_no == mat_no)
        )
        mat = existing.scalars().first()
        if mat:
            mat.kalem_tipi = kt
            if birim:
                mat.birim = birim
            if desc:
                mat.description = desc
            mat.source = "mm03_import"
            updated += 1
        else:
            db.add(MaterialMaster(
                material_no=mat_no, kalem_tipi=kt, birim=birim,
                description=desc, source="mm03_import"
            ))
            imported += 1

    await db.commit()
    wb.close()
    return {"total": imported + updated, "imported": imported, "updated": updated}


async def reprocess_project(project: BomProject, db: AsyncSession) -> dict:
    result = await db.execute(
        select(BomItem).where(BomItem.project_id == project.id).order_by(BomItem.row_number)
    )
    items = list(result.scalars().all())

    # Collect material keys
    mat_keys = set()
    for item in items:
        for key in (item.malzeme_no, item.title, item.ana_malzeme):
            if key:
                mat_keys.add(key.strip())

    material_map = {}
    if mat_keys:
        r = await db.execute(select(MaterialMaster).where(MaterialMaster.material_no.in_(mat_keys)))
        for m in r.scalars().all():
            material_map[m.material_no] = m

    level1_title = ""
    level2_title = ""
    parent_stack = []
    resolved = 0
    total = len(items)

    for item in items:
        if item.level == 1:
            level1_title = item.title or ""
        if item.level == 2:
            level2_title = item.title or ""

        while parent_stack and parent_stack[-1][0] >= item.level:
            parent_stack.pop()
        parent_qty_product = 1.0
        for _, pqty in parent_stack:
            parent_qty_product *= pqty

        if item.kalem_tipi_source == "manual":
            if not item.needs_review:
                resolved += 1
            parent_stack.append((item.level, item.quantity or 1.0))
            continue

        mat = (material_map.get((item.malzeme_no or "").strip())
               or material_map.get((item.title or "").strip())
               or material_map.get((item.ana_malzeme or "").strip()))

        derived = apply_rules(
            level=item.level, title=item.title or "", malzeme_no=item.malzeme_no or "",
            ana_malzeme=item.ana_malzeme or "", sap_usage=item.sap_usage or "",
            quantity=item.quantity or 1.0, level1_title=level1_title,
            level2_title=level2_title, material_kalem_tipi=mat.kalem_tipi if mat else None,
            material_birim=mat.birim if mat else None, parent_qty_product=parent_qty_product,
            parent_montaj_kt=None,
        )

        for k, v in derived.items():
            setattr(item, k, v)
        item.birlestirme = f"{derived['montaj']}|{item.title}|{derived['malzeme_no_sap']}"

        if not item.needs_review:
            resolved += 1

        parent_stack.append((item.level, item.quantity or 1.0))

    project.resolved_rows = resolved
    project.unresolved_rows = total - resolved
    project.status = "completed" if project.unresolved_rows == 0 else "review"
    await db.commit()

    return {"total": total, "resolved": resolved}


async def export_master_excel(project_id: int, db: AsyncSession) -> BytesIO:
    result = await db.execute(
        select(BomItem).where(BomItem.project_id == project_id).order_by(BomItem.row_number)
    )
    items = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Master BOM"

    headers = PLM_COLS + [
        "Uzmanlik", "Montaj", "MalzemeNo_SAP", "AnaMalzeme_Derived",
        "Birlestirme", "KalemTipi", "Siparis", "Dagitim", "Birim",
        "ToplamMiktar", "KalemTipiSource", "NeedsReview",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        plm_vals = [
            item.level, item.title, item.revision, item.quantity, item.description,
            item.maturity_state, item.owner, item.catia_aciklama, item.yedek_parca_mi,
            item.malzeme_standart_dokumani, item.tolerans_dokumani, item.dokuman_malzeme_turu,
            item.malzeme_no, item.istatistiksel_proses_kontrol, item.parca_standart_dokumani,
            item.sap_usage, item.yanmazlik_parametresi, item.hacim, item.boya_kodu,
            item.yuzey_alani, item.homologasyon_dokumani, item.emisyon_faktoru,
            item.finishing_standart_dokumani, item.referans_resim, item.kutle,
            item.sertlik, item.proje_kodu, item.kullanim_miktari, item.ana_malzeme_grubu,
            item.isil_islem_dokumani, item.ana_malzeme,
        ]
        derived_vals = [
            item.uzmanlik, item.montaj, item.malzeme_no_sap, item.ana_malzeme_derived,
            item.birlestirme, item.kalem_tipi, item.siparis, item.dagitim, item.birim,
            item.toplam_miktar, item.kalem_tipi_source, "EVET" if item.needs_review else "HAYIR",
        ]
        all_vals = plm_vals + derived_vals

        fill = LEVEL_FILLS.get(item.level)
        for col_idx, val in enumerate(all_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if item.level <= 1:
                cell.font = Font(bold=True, size=11)
            elif item.level == 2:
                cell.font = Font(bold=True, size=10)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
